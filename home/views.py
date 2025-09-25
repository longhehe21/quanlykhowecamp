# home/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import KyTonKho, HangHoa, NhapHangHoa, TonKhoHangHoa, CongThucMon, ChiTietCongThucMon, XuatMonTheoFabi, TongHopXuatNguyenLieu, TonKhoLeTan
from .forms import HangHoaForm, NhapHangHoaForm, TonKhoHangHoaForm, TonKhoHangHoaImportForm, TonKhoHangHoaFilterForm, CongThucMonForm, ChiTietCongThucMonForm, XuatMonTheoFabiForm, TonKhoLeTanForm
from django.contrib import messages
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Sum, F
import openpyxl
from datetime import datetime, timedelta, date
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from django.http import HttpResponse ,JsonResponse
from django.views.decorators.http import require_GET

def tinh_tong_ton_kho(ngay):
    tong_ton_list = []
    hang_hoa_list = HangHoa.objects.all()

    for hang_hoa in hang_hoa_list:
        ton_kho = TonKhoHangHoa.objects.filter(
            hang_hoa=hang_hoa,
            ngay_ton=ngay
        ).first()
        ton_kho_value = ton_kho.ton_cuoi_ngay if ton_kho else 0

        ton_kho_le_tan = TonKhoLeTan.objects.filter(
            hang_hoa=hang_hoa,
            ngay_ton=ngay
        ).first()
        ton_kho_le_tan_value = ton_kho_le_tan.ton_cuoi_ngay if ton_kho_le_tan else 0

        ty_le_doi = hang_hoa.dinh_luong or 1
        quy_doi = ton_kho_le_tan_value * ty_le_doi
        tong_ton = ton_kho_value + quy_doi

        tong_ton_list.append({
            'ten_hang_hoa': hang_hoa.ten_hang_hoa,
            'don_vi': hang_hoa.don_vi_nguyen_lieu or 'Không xác định',  # Đơn vị nguyên liệu cho tổng tồn
            'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa or 'Không xác định',  # Đơn vị hàng hóa cho tồn kho lễ tân
            'ton_kho': ton_kho_value,
            'ton_kho_le_tan': ton_kho_le_tan_value,
            'quy_doi': quy_doi,
            'tong_ton': tong_ton,
            'dinh_luong': ty_le_doi
        })
    return tong_ton_list
def tinh_ton_kho(hang_hoa_id, ngay_ton):
    """
    Tính tồn đầu ngày và tồn cuối ngày cho một hàng hóa vào một ngày cụ thể.
    - Tồn đầu ngày = Tồn cuối ngày hôm trước + Nhập hàng hôm nay
    - Tồn cuối ngày = Tồn đầu ngày - Xuất hàng hôm nay
    """
    from django.db.models import Sum
    ngay_truoc = ngay_ton - timezone.timedelta(days=1)
    ton_cuoi_ngay_truoc = TonKhoHangHoa.objects.filter(
        hang_hoa_id=hang_hoa_id,
        ngay_ton=ngay_truoc
    ).order_by('-ngay_ton').first()
    ton_cuoi_ngay_truoc_value = ton_cuoi_ngay_truoc.ton_cuoi_ngay if ton_cuoi_ngay_truoc else 0.0

    nhap_hang_hoa = NhapHangHoa.objects.filter(
        hang_hoa_id=hang_hoa_id,
        ngay_nhap=ngay_ton
    ).aggregate(total_so_luong=Sum('so_luong'))['total_so_luong'] or 0.0

    ton_dau_ngay = ton_cuoi_ngay_truoc_value + nhap_hang_hoa

    xuat_hang_hoa = TongHopXuatNguyenLieu.objects.filter(
        hang_hoa_id=hang_hoa_id,
        ngay_xuat=ngay_ton
    ).aggregate(total_nguyen_lieu=Sum('nguyen_lieu_da_xuat'))['total_nguyen_lieu'] or 0.0

    ton_cuoi_ngay = ton_dau_ngay - xuat_hang_hoa
    return ton_dau_ngay, ton_cuoi_ngay

def tinh_ton_kho_le_tan(hang_hoa_id, ngay_ton):
    """Tính tồn đầu ngày cho kho lễ tân dựa trên tồn cuối ngày trước đó."""
    ngay_truoc = ngay_ton - timezone.timedelta(days=1)
    ton_cuoi_ngay_truoc = TonKhoLeTan.objects.filter(
        hang_hoa_id=hang_hoa_id,
        ngay_ton=ngay_truoc
    ).order_by('-ngay_ton').first()
    return ton_cuoi_ngay_truoc.ton_cuoi_ngay if ton_cuoi_ngay_truoc else 0.0

def tinh_luong_dung_tong_hop(hang_hoa_id, date_from, date_to):
    """
    Tính tổng lượng dùng hàng hóa từ ngày bắt đầu đến ngày kết thúc.
    """
    luong_dung_tong = 0.0
    current_date = date_from
    while current_date <= date_to:
        ton_kho = TonKhoHangHoa.objects.filter(
            hang_hoa_id=hang_hoa_id,
            ngay_ton=current_date
        ).first()
        if ton_kho:
            luong_dung_ngay = ton_kho.ton_dau_ngay - ton_kho.ton_cuoi_ngay
            luong_dung_tong += luong_dung_ngay
        current_date += timedelta(days=1)
    return luong_dung_tong


def tinh_ton_kho_le_tan(hang_hoa_id, ngay_ton):
    """Tính tồn đầu ngày cho kho lễ tân dựa trên tồn cuối ngày trước đó."""
    ngay_truoc = ngay_ton - timezone.timedelta(days=1)
    ton_cuoi_ngay_truoc = TonKhoLeTan.objects.filter(
        hang_hoa_id=hang_hoa_id,
        ngay_ton=ngay_truoc
    ).order_by('-ngay_ton').first()
    return ton_cuoi_ngay_truoc.ton_cuoi_ngay if ton_cuoi_ngay_truoc else 0.0

def quan_ly_hang_hoa(request):
    """View quản lý hàng hóa với xử lý bulk tồn kho"""
    if request.method == 'POST':
        # Xử lý bulk tồn kho hàng hóa
        if 'bulk_action' in request.POST and request.POST['bulk_action'] == 'save_ton_kho':
            ngay_ton_str = request.POST.get('ngay_ton')
            try:
                ngay_ton = datetime.strptime(ngay_ton_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Ngày tồn không hợp lệ.'}, status=400)
            
            success_count = 0
            error_count = 0
            messages_list = []

            # Lấy tất cả items từ form
            for key in request.POST:
                if key.startswith('ton_kho_items['):
                    try:
                        index_str = key.split('[')[1].split(']')[0]
                        index = int(index_str)

                        hang_hoa_id = request.POST.get(f'ton_kho_items[{index}][hang_hoa_id]', None)  # Cho phép None
                        ton_dau_ngay_input = request.POST.get(f'ton_kho_items[{index}][ton_dau_ngay]', None)
                        ton_cuoi_ngay_input = request.POST.get(f'ton_kho_items[{index}][ton_cuoi_ngay]', '0')  # Mặc định 0
                        don_vi = request.POST.get(f'ton_kho_items[{index}][don_vi_nguyen_lieu]', '')

                        # Nếu hang_hoa_id rỗng, bỏ qua dòng này
                        if not hang_hoa_id:
                            messages_list.append(f"Dòng {index + 1}: Thiếu ID hàng hóa, bỏ qua dòng này")
                            continue

                        try:
                            hang_hoa = HangHoa.objects.get(id=hang_hoa_id)
                        except HangHoa.DoesNotExist:
                            error_count += 1
                            messages_list.append(f"Dòng {index + 1}: Hàng hóa ID {hang_hoa_id} không tồn tại")
                            continue

                        # Chuyển đổi ton_cuoi_ngay, mặc định 0 nếu không hợp lệ
                        try:
                            ton_cuoi_ngay = float(ton_cuoi_ngay_input)
                            if ton_cuoi_ngay < 0:
                                ton_cuoi_ngay = 0.0  # Đặt lại về 0 nếu âm
                                messages_list.append(f"Dòng {index + 1}: Tồn cuối ngày < 0, gán = 0")
                        except ValueError:
                            ton_cuoi_ngay = 0.0  # Mặc định 0 nếu không hợp lệ
                            messages_list.append(f"Dòng {index + 1}: Tồn cuối ngày không hợp lệ, gán = 0")

                        # Tính lại ton_dau_ngay từ logic
                        ton_dau_tinh_toan, _ = tinh_ton_kho(hang_hoa.id, ngay_ton)

                        # Cập nhật hoặc tạo mới
                        ton_kho, created = TonKhoHangHoa.objects.update_or_create(
                            hang_hoa=hang_hoa,
                            ngay_ton=ngay_ton,
                            defaults={
                                'ton_dau_ngay': ton_dau_tinh_toan,
                                'ton_cuoi_ngay': ton_cuoi_ngay,
                                'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa or don_vi
                            }
                        )

                        if created:
                            success_count += 1
                            messages_list.append(f"Tạo mới: {hang_hoa.ten_hang_hoa}")
                        else:
                            success_count += 1
                            messages_list.append(f"Cập nhật: {hang_hoa.ten_hang_hoa}")

                    except Exception as e:
                        error_count += 1
                        messages_list.append(f"Dòng {index + 1}: Lỗi không xác định: {str(e)}")

            # Trả về JSON response
            response_data = {
                'success': success_count > 0,
                'message': f'Đã lưu thành công {success_count} bản ghi, {error_count} lỗi' if success_count > 0 or error_count > 0 else 'Không có thay đổi nào được thực hiện',
                'messages_list': messages_list[:5]  # Giới hạn 5 tin nhắn
            }
            return JsonResponse(response_data)

        # Xử lý bulk nhập hàng hóa
        if 'bulk_action' in request.POST and request.POST['bulk_action'] == 'save_nhap_hang':
            ngay_nhap_str = request.POST.get('ngay_nhap')
            try:
                ngay_nhap = datetime.strptime(ngay_nhap_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Ngày nhập không hợp lệ.'}, status=400)
            
            success_count = 0
            error_count = 0
            messages_list = []

            for key in request.POST:
                if key.startswith('nhap_hang_items['):
                    try:
                        index_str = key.split('[')[1].split(']')[0]
                        index = int(index_str)

                        hang_hoa_id = request.POST.get(f'nhap_hang_items[{index}][hang_hoa_id]', None)
                        so_luong_input = request.POST.get(f'nhap_hang_items[{index}][so_luong]', '0')
                        don_vi = request.POST.get(f'nhap_hang_items[{index}][don_vi_hang_hoa]', '')

                        if not hang_hoa_id:
                            messages_list.append(f"Dòng {index + 1}: Thiếu ID hàng hóa, bỏ qua dòng này")
                            continue

                        try:
                            hang_hoa = HangHoa.objects.get(id=hang_hoa_id)
                        except HangHoa.DoesNotExist:
                            error_count += 1
                            messages_list.append(f"Dòng {index + 1}: Hàng hóa ID {hang_hoa_id} không tồn tại")
                            continue

                        try:
                            so_luong = float(so_luong_input)
                            if so_luong < 0:
                                so_luong = 0.0
                                messages_list.append(f"Dòng {index + 1}: Số lượng < 0, gán = 0")
                        except ValueError:
                            so_luong = 0.0
                            messages_list.append(f"Dòng {index + 1}: Số lượng không hợp lệ, gán = 0")

                        if so_luong > 0:  # Chỉ lưu nếu số lượng > 0
                            nhap_hang, created = NhapHangHoa.objects.update_or_create(
                                hang_hoa=hang_hoa,
                                ngay_nhap=ngay_nhap,
                                defaults={
                                    'so_luong': so_luong,
                                    'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa or don_vi
                                }
                            )

                            # Cập nhật tồn kho hàng hóa
                            ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(hang_hoa.id, ngay_nhap)
                            TonKhoHangHoa.objects.update_or_create(
                                hang_hoa=hang_hoa,
                                ngay_ton=ngay_nhap,
                                defaults={
                                    'ton_dau_ngay': ton_dau_ngay,
                                    'ton_cuoi_ngay': ton_cuoi_ngay + so_luong,  # Cập nhật tồn cuối
                                    'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa
                                }
                            )

                            if created:
                                success_count += 1
                                messages_list.append(f"Tạo mới: {hang_hoa.ten_hang_hoa}")
                            else:
                                success_count += 1
                                messages_list.append(f"Cập nhật: {hang_hoa.ten_hang_hoa}")

                    except Exception as e:
                        error_count += 1
                        messages_list.append(f"Dòng {index + 1}: Lỗi không xác định: {str(e)}")

            response_data = {
                'success': success_count > 0,
                'message': f'Đã lưu thành công {success_count} bản ghi, {error_count} lỗi',
                'messages_list': messages_list[:5]
            }
            return JsonResponse(response_data)

        # Xử lý bulk tồn kho lễ tân
        if 'bulk_action' in request.POST and request.POST['bulk_action'] == 'save_ton_kho_le_tan':
            ngay_ton_str = request.POST.get('ngay_ton')
            try:
                ngay_ton = datetime.strptime(ngay_ton_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Ngày tồn không hợp lệ.'}, status=400)
            
            success_count = 0
            error_count = 0
            messages_list = []

            for key in request.POST:
                if key.startswith('ton_kho_le_tan_items['):
                    try:
                        index_str = key.split('[')[1].split(']')[0]
                        index = int(index_str)

                        hang_hoa_id = request.POST.get(f'ton_kho_le_tan_items[{index}][hang_hoa_id]', None)
                        ton_dau_ngay_input = request.POST.get(f'ton_kho_le_tan_items[{index}][ton_dau_ngay]', None)
                        ton_cuoi_ngay_input = request.POST.get(f'ton_kho_le_tan_items[{index}][ton_cuoi_ngay]', '0')
                        don_vi = request.POST.get(f'ton_kho_le_tan_items[{index}][don_vi_hang_hoa]', '')

                        if not hang_hoa_id:
                            messages_list.append(f"Dòng {index + 1}: Thiếu ID hàng hóa, bỏ qua dòng này")
                            continue

                        try:
                            hang_hoa = HangHoa.objects.get(id=hang_hoa_id)
                        except HangHoa.DoesNotExist:
                            error_count += 1
                            messages_list.append(f"Dòng {index + 1}: Hàng hóa ID {hang_hoa_id} không tồn tại")
                            continue

                        try:
                            ton_cuoi_ngay = float(ton_cuoi_ngay_input)
                            if ton_cuoi_ngay < 0:
                                ton_cuoi_ngay = 0.0
                                messages_list.append(f"Dòng {index + 1}: Tồn cuối ngày < 0, gán = 0")
                        except ValueError:
                            ton_cuoi_ngay = 0.0
                            messages_list.append(f"Dòng {index + 1}: Tồn cuối ngày không hợp lệ, gán = 0")

                        # Tính lại ton_dau_ngay từ logic
                        ton_dau_ngay = tinh_ton_kho_le_tan(hang_hoa.id, ngay_ton)

                        ton_kho_le_tan, created = TonKhoLeTan.objects.update_or_create(
                            hang_hoa=hang_hoa,
                            ngay_ton=ngay_ton,
                            defaults={
                                'ton_dau_ngay': ton_dau_ngay,
                                'ton_cuoi_ngay': ton_cuoi_ngay,
                                'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa or don_vi
                            }
                        )

                        if created:
                            success_count += 1
                            messages_list.append(f"Tạo mới: {hang_hoa.ten_hang_hoa}")
                        else:
                            success_count += 1
                            messages_list.append(f"Cập nhật: {hang_hoa.ten_hang_hoa}")

                    except Exception as e:
                        error_count += 1
                        messages_list.append(f"Dòng {index + 1}: Lỗi không xác định: {str(e)}")

            response_data = {
                'success': success_count > 0,
                'message': f'Đã lưu thành công {success_count} bản ghi, {error_count} lỗi',
                'messages_list': messages_list[:5]
            }
            return JsonResponse(response_data)

        # Xử lý import Excel
        if 'excel_file' in request.FILES:
            import_form = TonKhoHangHoaImportForm(request.POST, request.FILES)
            if import_form.is_valid():
                excel_file = request.FILES['excel_file']
                import_type = request.POST.get('import_type')

                try:
                    df = pd.read_excel(excel_file)
                    if df.empty:
                        messages.error(request, 'File Excel không chứa dữ liệu.')
                        return redirect('quan_ly_hang_hoa')

                    if import_type == 'ton_kho':
                        required_columns = ['hang_hoa', 'ngay_ton', 'ton_cuoi_ngay']
                        if not all(col in df.columns for col in required_columns):
                            messages.error(request, f'File thiếu cột bắt buộc. Yêu cầu: {", ".join(required_columns)}.')
                            return redirect('quan_ly_hang_hoa')

                        if 'ngay_ton' in df.columns:
                            df['ngay_ton'] = pd.to_datetime(df['ngay_ton'], errors='coerce').dt.date
                            df = df.dropna(subset=['ngay_ton'])

                        for index, row in df.iterrows():
                            ten_hang_hoa = row.get('hang_hoa')
                            ngay_ton = row.get('ngay_ton')
                            ton_cuoi_ngay = row.get('ton_cuoi_ngay')

                            if not all([ten_hang_hoa, ngay_ton, ton_cuoi_ngay is not None]):
                                messages.error(request, f'Dòng {index + 2}: Thiếu dữ liệu bắt buộc (hang_hoa, ngay_ton, ton_cuoi_ngay).')
                                continue

                            try:
                                if pd.isna(ngay_ton):
                                    raise ValueError("Ngày không hợp lệ")
                                ngay_ton = ngay_ton

                                hang_hoa = HangHoa.objects.filter(ten_hang_hoa=ten_hang_hoa).first()
                                if not hang_hoa:
                                    messages.error(request, f'Dòng {index + 2}: Hàng hóa "{ten_hang_hoa}" không tồn tại.')
                                    continue

                                ton_dau_ngay, _ = tinh_ton_kho(hang_hoa.id, ngay_ton)
                                ton_cuoi_ngay = float(ton_cuoi_ngay) if pd.notna(ton_cuoi_ngay) else ton_dau_ngay

                                TonKhoHangHoa.objects.update_or_create(
                                    hang_hoa=hang_hoa,
                                    ngay_ton=ngay_ton,
                                    defaults={
                                        'ton_dau_ngay': ton_dau_ngay,
                                        'ton_cuoi_ngay': ton_cuoi_ngay,
                                        'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa
                                    }
                                )
                            except ValueError as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi định dạng ngày hoặc số: {str(e)}')
                                continue
                            except Exception as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi không xác định: {str(e)}')
                                continue
                        messages.success(request, 'Nhập dữ liệu tồn kho từ Excel thành công!')

                    elif import_type == 'nhap_hang_hoa':
                        required_columns = ['hang_hoa', 'ngay_nhap', 'don_vi_hang_hoa', 'so_luong']
                        if not all(col in df.columns for col in required_columns):
                            messages.error(request, f'File thiếu cột bắt buộc. Yêu cầu: {", ".join(required_columns)}.')
                            return redirect('quan_ly_hang_hoa')

                        if 'ngay_nhap' in df.columns:
                            df['ngay_nhap'] = pd.to_datetime(df['ngay_nhap'], errors='coerce').dt.date
                            df = df.dropna(subset=['ngay_nhap'])
                        else:
                            messages.error(request, 'Cột "ngay_nhap" không tồn tại trong file Excel.')
                            return redirect('quan_ly_hang_hoa')

                        for index, row in df.iterrows():
                            ten_hang_hoa = row.get('hang_hoa')
                            ngay_nhap = row.get('ngay_nhap')
                            don_vi_hang_hoa = row.get('don_vi_hang_hoa')
                            so_luong = row.get('so_luong')

                            if not all([ten_hang_hoa, ngay_nhap, don_vi_hang_hoa, so_luong is not None]):
                                messages.error(request, f'Dòng {index + 2}: Thiếu dữ liệu bắt buộc (hang_hoa, ngay_nhap, don_vi_hang_hoa, so_luong).')
                                continue

                            try:
                                if pd.isna(ngay_nhap):
                                    raise ValueError("Ngày không hợp lệ!")
                                ngay_nhap = ngay_nhap

                                hang_hoa = HangHoa.objects.filter(ten_hang_hoa=ten_hang_hoa).first()
                                if not hang_hoa:
                                    messages.error(request, f'Dòng {index + 2}: Hàng hóa "{ten_hang_hoa}" không tồn tại.')
                                    continue

                                so_luong = float(so_luong) if pd.notna(so_luong) else 0
                                if so_luong <= 0:
                                    messages.error(request, f'Dòng {index + 2}: Số lượng phải lớn hơn 0.')
                                    continue

                                NhapHangHoa.objects.update_or_create(
                                    hang_hoa=hang_hoa,
                                    ngay_nhap=ngay_nhap,
                                    defaults={
                                        'don_vi_hang_hoa': don_vi_hang_hoa,
                                        'so_luong': so_luong
                                    }
                                )

                                ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(hang_hoa.id, ngay_nhap)
                                existing_ton_kho = TonKhoHangHoa.objects.filter(
                                    hang_hoa=hang_hoa,
                                    ngay_ton=ngay_nhap
                                ).first()
                                if existing_ton_kho:
                                    existing_ton_kho.ton_dau_ngay = ton_dau_ngay
                                    existing_ton_kho.ton_cuoi_ngay = ton_cuoi_ngay + so_luong
                                    existing_ton_kho.save()
                                else:
                                    TonKhoHangHoa.objects.create(
                                        hang_hoa=hang_hoa,
                                        ngay_ton=ngay_nhap,
                                        ton_dau_ngay=ton_dau_ngay,
                                        ton_cuoi_ngay=ton_dau_ngay + so_luong,
                                        don_vi_hang_hoa=don_vi_hang_hoa
                                    )
                            except ValueError as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi định dạng ngày hoặc số: {str(e)}')
                                continue
                            except Exception as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi không xác định: {str(e)}')
                                continue
                        messages.success(request, 'Nhập dữ liệu hàng hóa từ Excel thành công!')

                    elif import_type == 'ton_kho_le_tan':
                        required_columns = ['hang_hoa', 'ngay_ton', 'ton_cuoi_ngay']
                        if not all(col in df.columns for col in required_columns):
                            messages.error(request, f'File thiếu cột bắt buộc. Yêu cầu: {", ".join(required_columns)}.')
                            return redirect('quan_ly_hang_hoa')

                        if 'ngay_ton' in df.columns:
                            df['ngay_ton'] = pd.to_datetime(df['ngay_ton'], errors='coerce').dt.date
                            df = df.dropna(subset=['ngay_ton'])

                        for index, row in df.iterrows():
                            ten_hang_hoa = row.get('hang_hoa')
                            ngay_ton = row.get('ngay_ton')
                            ton_cuoi_ngay = row.get('ton_cuoi_ngay')

                            if not all([ten_hang_hoa, ngay_ton, ton_cuoi_ngay is not None]):
                                messages.error(request, f'Dòng {index + 2}: Thiếu dữ liệu bắt buộc (hang_hoa, ngay_ton, ton_cuoi_ngay).')
                                continue

                            try:
                                if pd.isna(ngay_ton):
                                    raise ValueError("Ngày không hợp lệ")
                                ngay_ton = ngay_ton

                                hang_hoa = HangHoa.objects.filter(ten_hang_hoa=ten_hang_hoa).first()
                                if not hang_hoa:
                                    messages.error(request, f'Dòng {index + 2}: Hàng hóa "{ten_hang_hoa}" không tồn tại.')
                                    continue

                                ton_dau_ngay = tinh_ton_kho_le_tan(hang_hoa.id, ngay_ton)
                                ton_cuoi_ngay = float(ton_cuoi_ngay) if pd.notna(ton_cuoi_ngay) else ton_dau_ngay

                                TonKhoLeTan.objects.update_or_create(
                                    hang_hoa=hang_hoa,
                                    ngay_ton=ngay_ton,
                                    defaults={
                                        'ton_dau_ngay': ton_dau_ngay,
                                        'ton_cuoi_ngay': ton_cuoi_ngay,
                                        'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa
                                    }
                                )
                            except ValueError as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi định dạng ngày hoặc số: {str(e)}')
                                continue
                            except Exception as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi không xác định: {str(e)}')
                                continue
                        messages.success(request, 'Nhập dữ liệu tồn kho lễ tân từ Excel thành công!')

                    else:
                        messages.error(request, 'Loại nhập không hợp lệ.')
                        return redirect('quan_ly_hang_hoa')

                except Exception as e:
                    messages.error(request, f'Lỗi khi đọc file Excel: {str(e)}')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, 'Lỗi khi nhập file Excel. Vui lòng kiểm tra định dạng.')
        
        # Xử lý form nhập hàng hóa
        elif 'nhap_hang_hoa_form' in request.POST:
            nhap_hang_hoa_form = NhapHangHoaForm(request.POST)
            if nhap_hang_hoa_form.is_valid():
                nhap_hang_hoa = nhap_hang_hoa_form.save(commit=False)
                if nhap_hang_hoa.so_luong <= 0:
                    messages.error(request, 'Số lượng phải lớn hơn 0.')
                    return redirect('quan_ly_hang_hoa')
                nhap_hang_hoa.don_vi_hang_hoa = nhap_hang_hoa.hang_hoa.don_vi_hang_hoa
                nhap_hang_hoa.save()
                existing_ton_kho = TonKhoHangHoa.objects.filter(
                    hang_hoa=nhap_hang_hoa.hang_hoa,
                    ngay_ton=nhap_hang_hoa.ngay_nhap
                ).first()
                ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(nhap_hang_hoa.hang_hoa.id, nhap_hang_hoa.ngay_nhap)
                if existing_ton_kho:
                    existing_ton_kho.ton_dau_ngay = ton_dau_ngay
                    existing_ton_kho.ton_cuoi_ngay = ton_cuoi_ngay + nhap_hang_hoa.so_luong
                    existing_ton_kho.save()
                else:
                    TonKhoHangHoa.objects.create(
                        hang_hoa=nhap_hang_hoa.hang_hoa,
                        ngay_ton=nhap_hang_hoa.ngay_nhap,
                        ton_dau_ngay=ton_dau_ngay,
                        ton_cuoi_ngay=ton_dau_ngay + nhap_hang_hoa.so_luong,
                        don_vi_hang_hoa=nhap_hang_hoa.hang_hoa.don_vi_hang_hoa
                    )
                messages.success(request, 'Nhập hàng hóa thành công!')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, 'Lỗi khi nhập hàng hóa. Vui lòng kiểm tra lại.')
        
        # Xử lý form tồn kho lễ tân
        elif 'ton_kho_le_tan_form' in request.POST:
            ton_kho_le_tan_form = TonKhoLeTanForm(request.POST)
            if ton_kho_le_tan_form.is_valid():
                ton_kho = ton_kho_le_tan_form.save(commit=False)
                ton_dau_ngay = tinh_ton_kho_le_tan(ton_kho.hang_hoa.id, ton_kho.ngay_ton)
                ton_kho.ton_dau_ngay = ton_dau_ngay
                ton_kho.ton_cuoi_ngay = ton_kho_le_tan_form.cleaned_data['ton_cuoi_ngay']
                ton_kho.don_vi_hang_hoa = ton_kho.hang_hoa.don_vi_hang_hoa
                ton_kho.save()
                messages.success(request, 'Thêm tồn kho lễ tân thành công!')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, f'Lỗi khi thêm tồn kho lễ tân: {ton_kho_le_tan_form.errors.as_text()}')
        
        # Xử lý form tồn kho hàng hóa
        else:
            ton_kho_form = TonKhoHangHoaForm(request.POST)
            if ton_kho_form.is_valid():
                ton_kho = ton_kho_form.save(commit=False)
                ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(ton_kho.hang_hoa.id, ton_kho.ngay_ton)
                ton_kho.ton_dau_ngay = ton_dau_ngay
                ton_kho.ton_cuoi_ngay = ton_kho_form.cleaned_data['ton_cuoi_ngay']
                ton_kho.don_vi_hang_hoa = ton_kho.hang_hoa.don_vi_hang_hoa
                ton_kho.save()
                messages.success(request, 'Thêm tồn kho thành công!')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, 'Lỗi khi thêm tồn kho. Vui lòng kiểm tra lại.')
    # Phần còn lại của view (không thay đổi)
    ton_kho_form = TonKhoHangHoaForm()
    import_form = TonKhoHangHoaImportForm()
    nhap_hang_hoa_form = NhapHangHoaForm()
    filter_form = TonKhoHangHoaFilterForm(request.GET or None)
    ton_kho_le_tan_form = TonKhoLeTanForm()
    filter_le_tan_form = TonKhoHangHoaFilterForm(request.GET or None, prefix='le_tan')

    ton_kho_list = TonKhoHangHoa.objects.all()
    nhap_hang_hoa_list = NhapHangHoa.objects.filter(hang_hoa__isnull=False).order_by('-ngay_nhap')
    ton_kho_le_tan_list = TonKhoLeTan.objects.all()

    # Lọc tồn kho hàng hóa
    search_query = request.GET.get('search', '')
    if search_query:
        ton_kho_list = ton_kho_list.filter(hang_hoa__ten_hang_hoa__icontains=search_query)

    if filter_form.is_valid():
        ngay_bat_dau = filter_form.cleaned_data.get('ngay_bat_dau')
        ngay_ket_thuc = filter_form.cleaned_data.get('ngay_ket_thuc')
        if ngay_bat_dau:
            ton_kho_list = ton_kho_list.filter(ngay_ton__gte=ngay_bat_dau)
        if ngay_ket_thuc:
            ton_kho_list = ton_kho_list.filter(ngay_ton__lte=ngay_ket_thuc)

    # Lọc tồn kho lễ tân
    search_le_tan_query = request.GET.get('search_le_tan', '')
    if search_le_tan_query:
        ton_kho_le_tan_list = ton_kho_le_tan_list.filter(hang_hoa__ten_hang_hoa__icontains=search_le_tan_query)

    if filter_le_tan_form.is_valid():
        ngay_bat_dau = filter_le_tan_form.cleaned_data.get('ngay_bat_dau')
        ngay_ket_thuc = filter_le_tan_form.cleaned_data.get('ngay_ket_thuc')
        if ngay_bat_dau:
            ton_kho_le_tan_list = ton_kho_le_tan_list.filter(ngay_ton__gte=ngay_bat_dau)
        if ngay_ket_thuc:
            ton_kho_le_tan_list = ton_kho_le_tan_list.filter(ngay_ton__lte=ngay_ket_thuc)

    # Lọc lịch sử nhập hàng theo ngày
    ngay_nhap_bat_dau = request.GET.get('ngay_nhap_bat_dau')
    ngay_nhap_ket_thuc = request.GET.get('ngay_nhap_ket_thuc')
    if ngay_nhap_bat_dau:
        try:
            ngay_nhap_bat_dau = pd.to_datetime(ngay_nhap_bat_dau).date()
            nhap_hang_hoa_list = nhap_hang_hoa_list.filter(ngay_nhap__gte=ngay_nhap_bat_dau)
        except ValueError:
            messages.error(request, 'Định dạng ngày nhập bắt đầu không hợp lệ.')
    if ngay_nhap_ket_thuc:
        try:
            ngay_nhap_ket_thuc = pd.to_datetime(ngay_nhap_ket_thuc).date()
            nhap_hang_hoa_list = nhap_hang_hoa_list.filter(ngay_nhap__lte=ngay_nhap_ket_thuc)
        except ValueError:
            messages.error(request, 'Định dạng ngày nhập kết thúc không hợp lệ.')

    ton_kho_list = ton_kho_list.order_by('ngay_ton')
    ton_kho_le_tan_list = ton_kho_le_tan_list.order_by('ngay_ton')

    hang_hoa_list = HangHoa.objects.all()
    hang_hoa_list_json = json.dumps(
        list(hang_hoa_list.values('pk', 'ten_hang_hoa', 'don_vi_hang_hoa', 'don_vi_nguyen_lieu')),
        cls=DjangoJSONEncoder
    )

    # Lọc ngày để tính lượng dùng
    date_from = request.GET.get('luong_dung_date_from')
    date_to = request.GET.get('luong_dung_date_to')
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = date(2025, 5, 14)
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = date.today()
        if date_from > date_to:
            messages.error(request, 'Ngày bắt đầu phải nhỏ hơn hoặc bằng ngày kết thúc.')
            date_from = date_to
    except ValueError:
        messages.error(request, 'Định dạng ngày không hợp lệ.')
        date_from = date(2025, 5, 14)
        date_to = date.today()

    # Tính lượng dùng cho từng hàng hóa
    luong_dung_list = []
    for hang_hoa in hang_hoa_list:
        luong_dung_tong = tinh_luong_dung_tong_hop(hang_hoa.id, date_from, date_to)
        luong_dung_list.append({
            'hang_hoa': hang_hoa,
            'luong_dung_tong': luong_dung_tong,
            'don_vi': hang_hoa.don_vi_hang_hoa
        })

    # Lọc ngày cho tổng tồn kho
    tong_ton_date = request.GET.get('tong_ton_date')
    try:
        if tong_ton_date:
            tong_ton_date = datetime.strptime(tong_ton_date, '%Y-%m-%d').date()
        else:
            tong_ton_date = date.today()
    except ValueError:
        messages.error(request, 'Định dạng ngày tổng tồn kho không hợp lệ.')
        tong_ton_date = date.today()

    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            import_form = TonKhoHangHoaImportForm(request.POST, request.FILES)
            if import_form.is_valid():
                excel_file = request.FILES['excel_file']
                import_type = request.POST.get('import_type')

                try:
                    df = pd.read_excel(excel_file)
                    if df.empty:
                        messages.error(request, 'File Excel không chứa dữ liệu.')
                        return redirect('quan_ly_hang_hoa')

                    if import_type == 'ton_kho':
                        required_columns = ['hang_hoa', 'ngay_ton', 'ton_cuoi_ngay']
                        if not all(col in df.columns for col in required_columns):
                            messages.error(request, f'File thiếu cột bắt buộc. Yêu cầu: {", ".join(required_columns)}.')
                            return redirect('quan_ly_hang_hoa')

                        if 'ngay_ton' in df.columns:
                            df['ngay_ton'] = pd.to_datetime(df['ngay_ton'], errors='coerce').dt.date
                            df = df.dropna(subset=['ngay_ton'])

                        for index, row in df.iterrows():
                            ten_hang_hoa = row.get('hang_hoa')
                            ngay_ton = row.get('ngay_ton')
                            ton_cuoi_ngay = row.get('ton_cuoi_ngay')

                            if not all([ten_hang_hoa, ngay_ton, ton_cuoi_ngay is not None]):
                                messages.error(request, f'Dòng {index + 2}: Thiếu dữ liệu bắt buộc (hang_hoa, ngay_ton, ton_cuoi_ngay).')
                                continue

                            try:
                                if pd.isna(ngay_ton):
                                    raise ValueError("Ngày không hợp lệ")
                                ngay_ton = ngay_ton

                                hang_hoa = HangHoa.objects.filter(ten_hang_hoa=ten_hang_hoa).first()
                                if not hang_hoa:
                                    messages.error(request, f'Dòng {index + 2}: Hàng hóa "{ten_hang_hoa}" không tồn tại.')
                                    continue

                                ton_dau_ngay, _ = tinh_ton_kho(hang_hoa.id, ngay_ton)
                                ton_cuoi_ngay = float(ton_cuoi_ngay) if pd.notna(ton_cuoi_ngay) else ton_dau_ngay

                                TonKhoHangHoa.objects.update_or_create(
                                    hang_hoa=hang_hoa,
                                    ngay_ton=ngay_ton,
                                    defaults={
                                        'ton_dau_ngay': ton_dau_ngay,
                                        'ton_cuoi_ngay': ton_cuoi_ngay,
                                        'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa
                                    }
                                )
                            except ValueError as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi định dạng ngày hoặc số: {str(e)}')
                                continue
                            except Exception as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi không xác định: {str(e)}')
                                continue
                        messages.success(request, 'Nhập dữ liệu tồn kho từ Excel thành công!')

                    elif import_type == 'nhap_hang_hoa':
                        required_columns = ['hang_hoa', 'ngay_nhap', 'don_vi_hang_hoa', 'so_luong']
                        if not all(col in df.columns for col in required_columns):
                            messages.error(request, f'File thiếu cột bắt buộc. Yêu cầu: {", ".join(required_columns)}.')
                            return redirect('quan_ly_hang_hoa')

                        if 'ngay_nhap' in df.columns:
                            df['ngay_nhap'] = pd.to_datetime(df['ngay_nhap'], errors='coerce').dt.date
                            df = df.dropna(subset=['ngay_nhap'])
                        else:
                            messages.error(request, 'Cột "ngay_nhap" không tồn tại trong file Excel.')
                            return redirect('quan_ly_hang_hoa')

                        for index, row in df.iterrows():
                            ten_hang_hoa = row.get('hang_hoa')
                            ngay_nhap = row.get('ngay_nhap')
                            don_vi_hang_hoa = row.get('don_vi_hang_hoa')
                            so_luong = row.get('so_luong')

                            if not all([ten_hang_hoa, ngay_nhap, don_vi_hang_hoa, so_luong is not None]):
                                messages.error(request, f'Dòng {index + 2}: Thiếu dữ liệu bắt buộc (hang_hoa, ngay_nhap, don_vi_hang_hoa, so_luong).')
                                continue

                            try:
                                if pd.isna(ngay_nhap):
                                    raise ValueError("Ngày không hợp lệ!")
                                ngay_nhap = ngay_nhap

                                hang_hoa = HangHoa.objects.filter(ten_hang_hoa=ten_hang_hoa).first()
                                if not hang_hoa:
                                    messages.error(request, f'Dòng {index + 2}: Hàng hóa "{ten_hang_hoa}" không tồn tại.')
                                    continue

                                so_luong = float(so_luong) if pd.notna(so_luong) else 0
                                if so_luong <= 0:
                                    messages.error(request, f'Dòng {index + 2}: Số lượng phải lớn hơn 0.')
                                    continue

                                NhapHangHoa.objects.update_or_create(
                                    hang_hoa=hang_hoa,
                                    ngay_nhap=ngay_nhap,
                                    defaults={
                                        'don_vi_hang_hoa': don_vi_hang_hoa,
                                        'so_luong': so_luong
                                    }
                                )

                                ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(hang_hoa.id, ngay_nhap)
                                existing_ton_kho = TonKhoHangHoa.objects.filter(
                                    hang_hoa=hang_hoa,
                                    ngay_ton=ngay_nhap
                                ).first()
                                if existing_ton_kho:
                                    existing_ton_kho.ton_dau_ngay = ton_dau_ngay
                                    existing_ton_kho.ton_cuoi_ngay = ton_cuoi_ngay + so_luong
                                    existing_ton_kho.save()
                                else:
                                    TonKhoHangHoa.objects.create(
                                        hang_hoa=hang_hoa,
                                        ngay_ton=ngay_nhap,
                                        ton_dau_ngay=ton_dau_ngay,
                                        ton_cuoi_ngay=ton_dau_ngay + so_luong,
                                        don_vi_hang_hoa=don_vi_hang_hoa
                                    )
                            except ValueError as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi định dạng ngày hoặc số: {str(e)}')
                                continue
                            except Exception as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi không xác định: {str(e)}')
                                continue
                        messages.success(request, 'Nhập dữ liệu hàng hóa từ Excel thành công!')

                    elif import_type == 'ton_kho_le_tan':
                        required_columns = ['hang_hoa', 'ngay_ton', 'ton_cuoi_ngay']
                        if not all(col in df.columns for col in required_columns):
                            messages.error(request, f'File thiếu cột bắt buộc. Yêu cầu: {", ".join(required_columns)}.')
                            return redirect('quan_ly_hang_hoa')

                        if 'ngay_ton' in df.columns:
                            df['ngay_ton'] = pd.to_datetime(df['ngay_ton'], errors='coerce').dt.date
                            df = df.dropna(subset=['ngay_ton'])

                        for index, row in df.iterrows():
                            ten_hang_hoa = row.get('hang_hoa')
                            ngay_ton = row.get('ngay_ton')
                            ton_cuoi_ngay = row.get('ton_cuoi_ngay')

                            if not all([ten_hang_hoa, ngay_ton, ton_cuoi_ngay is not None]):
                                messages.error(request, f'Dòng {index + 2}: Thiếu dữ liệu bắt buộc (hang_hoa, ngay_ton, ton_cuoi_ngay).')
                                continue

                            try:
                                if pd.isna(ngay_ton):
                                    raise ValueError("Ngày không hợp lệ")
                                ngay_ton = ngay_ton

                                hang_hoa = HangHoa.objects.filter(ten_hang_hoa=ten_hang_hoa).first()
                                if not hang_hoa:
                                    messages.error(request, f'Dòng {index + 2}: Hàng hóa "{ten_hang_hoa}" không tồn tại.')
                                    continue

                                ton_dau_ngay = tinh_ton_kho_le_tan(hang_hoa.id, ngay_ton)
                                ton_cuoi_ngay = float(ton_cuoi_ngay) if pd.notna(ton_cuoi_ngay) else ton_dau_ngay

                                TonKhoLeTan.objects.update_or_create(
                                    hang_hoa=hang_hoa,
                                    ngay_ton=ngay_ton,
                                    defaults={
                                        'ton_dau_ngay': ton_dau_ngay,
                                        'ton_cuoi_ngay': ton_cuoi_ngay,
                                        'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa
                                    }
                                )
                            except ValueError as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi định dạng ngày hoặc số: {str(e)}')
                                continue
                            except Exception as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi không xác định: {str(e)}')
                                continue
                        messages.success(request, 'Nhập dữ liệu tồn kho lễ tân từ Excel thành công!')

                    else:
                        messages.error(request, 'Loại nhập không hợp lệ.')
                        return redirect('quan_ly_hang_hoa')

                except Exception as e:
                    messages.error(request, f'Lỗi khi đọc file Excel: {str(e)}')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, 'Lỗi khi nhập file Excel. Vui lòng kiểm tra định dạng.')
        elif 'nhap_hang_hoa_form' in request.POST:
            nhap_hang_hoa_form = NhapHangHoaForm(request.POST)
            if nhap_hang_hoa_form.is_valid():
                nhap_hang_hoa = nhap_hang_hoa_form.save(commit=False)
                if nhap_hang_hoa.so_luong <= 0:
                    messages.error(request, 'Số lượng phải lớn hơn 0.')
                    return redirect('quan_ly_hang_hoa')
                nhap_hang_hoa.don_vi_hang_hoa = nhap_hang_hoa.hang_hoa.don_vi_hang_hoa
                nhap_hang_hoa.save()
                existing_ton_kho = TonKhoHangHoa.objects.filter(
                    hang_hoa=nhap_hang_hoa.hang_hoa,
                    ngay_ton=nhap_hang_hoa.ngay_nhap
                ).first()
                ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(nhap_hang_hoa.hang_hoa.id, nhap_hang_hoa.ngay_nhap)
                if existing_ton_kho:
                    existing_ton_kho.ton_dau_ngay = ton_dau_ngay
                    existing_ton_kho.ton_cuoi_ngay = ton_cuoi_ngay + nhap_hang_hoa.so_luong
                    existing_ton_kho.save()
                else:
                    TonKhoHangHoa.objects.create(
                        hang_hoa=nhap_hang_hoa.hang_hoa,
                        ngay_ton=nhap_hang_hoa.ngay_nhap,
                        ton_dau_ngay=ton_dau_ngay,
                        ton_cuoi_ngay=ton_dau_ngay + nhap_hang_hoa.so_luong,
                        don_vi_hang_hoa=nhap_hang_hoa.hang_hoa.don_vi_hang_hoa
                    )
                messages.success(request, 'Nhập hàng hóa thành công!')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, 'Lỗi khi nhập hàng hóa. Vui lòng kiểm tra lại.')
        elif 'ton_kho_le_tan_form' in request.POST:
            ton_kho_le_tan_form = TonKhoLeTanForm(request.POST)
            if ton_kho_le_tan_form.is_valid():
                ton_kho = ton_kho_le_tan_form.save(commit=False)
                ton_dau_ngay = tinh_ton_kho_le_tan(ton_kho.hang_hoa.id, ton_kho.ngay_ton)
                ton_kho.ton_dau_ngay = ton_dau_ngay
                ton_kho.ton_cuoi_ngay = ton_kho_le_tan_form.cleaned_data['ton_cuoi_ngay']
                ton_kho.don_vi_hang_hoa = ton_kho.hang_hoa.don_vi_hang_hoa
                ton_kho.save()
                messages.success(request, 'Thêm tồn kho lễ tân thành công!')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, f'Lỗi khi thêm tồn kho lễ tân: {ton_kho_le_tan_form.errors.as_text()}')
        else:
            ton_kho_form = TonKhoHangHoaForm(request.POST)
            if ton_kho_form.is_valid():
                ton_kho = ton_kho_form.save(commit=False)
                ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(ton_kho.hang_hoa.id, ton_kho.ngay_ton)
                ton_kho.ton_dau_ngay = ton_dau_ngay
                ton_kho.ton_cuoi_ngay = ton_kho_form.cleaned_data['ton_cuoi_ngay']
                ton_kho.don_vi_hang_hoa = ton_kho.hang_hoa.don_vi_hang_hoa
                ton_kho.save()
                messages.success(request, 'Thêm tồn kho thành công!')
                return redirect('quan_ly_hang_hoa')
            else:
                messages.error(request, 'Lỗi khi thêm tồn kho. Vui lòng kiểm tra lại.')

    tong_ton_kho_list = tinh_tong_ton_kho(tong_ton_date)

    context = {
        'ton_kho_form': ton_kho_form,
        'import_form': import_form,
        'nhap_hang_hoa_form': nhap_hang_hoa_form,
        'filter_form': filter_form,
        'ton_kho_le_tan_form': ton_kho_le_tan_form,
        'filter_le_tan_form': filter_le_tan_form,
        'ton_kho_list': ton_kho_list,
        'nhap_hang_hoa_list': nhap_hang_hoa_list,
        'ton_kho_le_tan_list': ton_kho_le_tan_list,
        'search_query': search_query,
        'search_le_tan_query': search_le_tan_query,
        'hang_hoa_list_json': hang_hoa_list_json,
        'ngay_nhap_bat_dau': ngay_nhap_bat_dau,
        'ngay_nhap_ket_thuc': ngay_nhap_ket_thuc,
        'luong_dung_list': luong_dung_list,
        'luong_dung_date_from': date_from,
        'luong_dung_date_to': date_to,
        'today': date.today(),
        'tong_ton_kho_list': tong_ton_kho_list,
        'tong_ton_date': tong_ton_date,
    }
    return render(request, 'home/quan_ly_hang_hoa.html', context)

def export_tong_ton_kho_excel(request):
    # Lấy tham số ngày từ request
    tong_ton_date = request.GET.get('tong_ton_date')
    try:
        tong_ton_date = datetime.strptime(tong_ton_date, '%Y-%m-%d').date() if tong_ton_date else None
    except ValueError:
        tong_ton_date = None

    # Lấy dữ liệu tổng tồn kho (sử dụng hàm tinh_tong_ton_kho để đồng bộ logic)
    tong_ton_kho_list = tinh_tong_ton_kho(tong_ton_date)

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Tong Ton Kho"

    # Tiêu đề cột
    headers = [
        'Tên Hàng Hóa',
        'Tồn Kho',
        'Đơn Vị Tồn Kho',
        'Tồn Kho Lễ Tân',
        'Đơn Vị Tồn Kho Lễ Tân',
        'Quy Đổi',
        'Tổng Tồn',
        'Đơn Vị Tổng Tồn',
        'Tỷ Lệ Quy Đổi'
    ]
    ws.append(headers)

    # Thêm dữ liệu
    for item in tong_ton_kho_list:
        ws.append([
            item['ten_hang_hoa'],
            item['ton_kho'],
            item['don_vi'],
            item['ton_kho_le_tan'],
            item['don_vi_hang_hoa'],
            item['quy_doi'],
            item['tong_ton'],
            item['don_vi'],
            item['dinh_luong']
        ])

    # Tự động điều chỉnh độ rộng cột
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True

    # Tạo response với file Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Tong_Ton_Kho_{tong_ton_date or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

def export_so_sanh_fabi_excel(request):
    # Lấy tham số tìm kiếm và lọc
    ten_hang_hoa = request.GET.get('ten_hang_hoa', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Chuyển đổi ngày nếu có
    try:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
    except ValueError:
        date_from = None
        date_to = None

    # Lấy dữ liệu từ TongHopXuatNguyenLieu (Nguyên Liệu Xuất - Fabi)
    tong_hop_query = TongHopXuatNguyenLieu.objects.all()
    if ten_hang_hoa:
        tong_hop_query = tong_hop_query.filter(hang_hoa__ten_hang_hoa__icontains=ten_hang_hoa)
    if date_from:
        tong_hop_query = tong_hop_query.filter(ngay_xuat__gte=date_from)
    if date_to:
        tong_hop_query = tong_hop_query.filter(ngay_xuat__lte=date_to)

    tong_hop_data = tong_hop_query.values('hang_hoa__ten_hang_hoa').annotate(
        tong_da_xuat=Sum('nguyen_lieu_da_xuat')
    )

    # Lấy dữ liệu lượng dùng từ TonKhoHangHoa
    luong_dung_query = TonKhoHangHoa.objects.all()
    if ten_hang_hoa:
        luong_dung_query = luong_dung_query.filter(hang_hoa__ten_hang_hoa__icontains=ten_hang_hoa)
    if date_from:
        luong_dung_query = luong_dung_query.filter(ngay_ton__gte=date_from)
    if date_to:
        luong_dung_query = luong_dung_query.filter(ngay_ton__lte=date_to)

    luong_dung_data = luong_dung_query.values('hang_hoa__ten_hang_hoa').annotate(
        luong_dung_tong=Sum(F('ton_dau_ngay') - F('ton_cuoi_ngay'))
    )

    # Kết hợp dữ liệu
    so_sanh_list = []
    hang_hoa_list = HangHoa.objects.all()
    if ten_hang_hoa:
        hang_hoa_list = hang_hoa_list.filter(ten_hang_hoa__icontains=ten_hang_hoa)

    for hang_hoa in hang_hoa_list:
        fabi_data = next((item for item in tong_hop_data if item['hang_hoa__ten_hang_hoa'] == hang_hoa.ten_hang_hoa), None)
        nguyen_lieu_da_xuat = fabi_data['tong_da_xuat'] if fabi_data else 0

        luong_dung = next((item for item in luong_dung_data if item['hang_hoa__ten_hang_hoa'] == hang_hoa.ten_hang_hoa), None)
        so_luong_ton = abs(luong_dung['luong_dung_tong']) if luong_dung and luong_dung['luong_dung_tong'] is not None else 0

        chenh_lech = so_luong_ton - nguyen_lieu_da_xuat

        if nguyen_lieu_da_xuat or so_luong_ton:
            so_sanh_list.append({
                'ten_hang_hoa': hang_hoa.ten_hang_hoa,
                'nguyen_lieu_da_xuat': nguyen_lieu_da_xuat,
                'so_luong_ton': so_luong_ton,
                'chenh_lech': chenh_lech,
                'don_vi': hang_hoa.don_vi_nguyen_lieu or 'Không xác định',
                'date_from': date_from or timezone.now().date(),
                'date_to': date_to or timezone.now().date(),
            })

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "So Sanh Fabi"

    # Tiêu đề cột
    headers = [
        'Tên Hàng Hóa',
        'Nguyên Liệu Xuất (Fabi)',
        'Lượng Dùng Thực Tế',
        'Chênh Lệch',
        'Đơn Vị',
        'Từ Ngày',
        'Đến Ngày'
    ]
    ws.append(headers)

    # Thêm dữ liệu
    for item in so_sanh_list:
        ws.append([
            item['ten_hang_hoa'],
            item['nguyen_lieu_da_xuat'],
            item['so_luong_ton'],
            item['chenh_lech'],
            item['don_vi'],
            item['date_from'].strftime('%d/%m/%Y'),
            item['date_to'].strftime('%d/%m/%Y')
        ])

    # Tự động điều chỉnh độ rộng cột
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True

    # Tạo response với file Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"So_Sanh_Fabi_{date_from or 'all'}_to_{date_to or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

def xuat_theo_mon(request):
    # Khởi tạo form ngay từ đầu để tránh lỗi UnboundLocalError
    cong_thuc_form = CongThucMonForm()
    chi_tiet_form = ChiTietCongThucMonForm()
    xuat_mon_form = XuatMonTheoFabiForm()

    if request.method == 'POST':
        if 'cong_thuc_mon_form' in request.POST:
            cong_thuc_form = CongThucMonForm(request.POST)
            if cong_thuc_form.is_valid():
                cong_thuc = cong_thuc_form.save()
                i = 0
                created_entries = 0
                while True:
                    hang_hoa_key = f'hang_hoa_{i}'
                    if hang_hoa_key not in request.POST:
                        break
                    hang_hoa_id = request.POST.get(hang_hoa_key)
                    dinh_luong_key = f'dinh_luong_{i}'
                    dinh_luong = request.POST.get(dinh_luong_key, 0)
                    if hang_hoa_id and dinh_luong:
                        try:
                            ChiTietCongThucMon.objects.create(
                                cong_thuc_mon=cong_thuc,
                                hang_hoa_id=hang_hoa_id,
                                dinh_luong=float(dinh_luong)
                            )
                            created_entries += 1
                        except Exception as e:
                            messages.error(request, f'Lỗi khi tạo chi tiết tại hàng {i+1}: {str(e)}')
                    i += 1
                if i == 0:
                    messages.warning(request, 'Bạn chưa thêm nguyên liệu nào cho công thức.')
                elif created_entries > 0:
                    messages.success(request, f'Thêm công thức thành công! Đã thêm {created_entries} nguyên liệu.')
                else:
                    messages.warning(request, 'Không có nguyên liệu nào được thêm do lỗi.')
                return redirect('xuat_theo_mon')
            else:
                messages.error(request, 'Lỗi khi thêm công thức. Vui lòng kiểm tra lại.')
        elif 'xuat_mon_fabi_form' in request.POST:
            xuat_mon_form = XuatMonTheoFabiForm(request.POST)
            if xuat_mon_form.is_valid():
                xuat_mon = xuat_mon_form.save()
                chi_tiet_list = ChiTietCongThucMon.objects.filter(cong_thuc_mon=xuat_mon.ten_mon)
                for chi_tiet in chi_tiet_list:
                    tong_hop, created = TongHopXuatNguyenLieu.objects.get_or_create(
                        cong_thuc_mon=xuat_mon.ten_mon,
                        hang_hoa=chi_tiet.hang_hoa,
                        ngay_xuat=xuat_mon.ngay_xuat,
                        defaults={
                            'dinh_luong': chi_tiet.dinh_luong,
                            'so_mon_xuat': xuat_mon.so_luong,
                            'nguyen_lieu_da_xuat': chi_tiet.dinh_luong * xuat_mon.so_luong
                        }
                    )
                    if not created:
                        tong_hop.so_mon_xuat += xuat_mon.so_luong
                        tong_hop.nguyen_lieu_da_xuat = chi_tiet.dinh_luong * tong_hop.so_mon_xuat
                        tong_hop.save()
                    if abs(tong_hop.nguyen_lieu_da_xuat - (tong_hop.dinh_luong * tong_hop.so_mon_xuat)) > 0.0001:
                        messages.error(request, f'Lỗi toàn vẹn dữ liệu cho {tong_hop.hang_hoa.ten_hang_hoa} ngày {tong_hop.ngay_xuat}')
                messages.success(request, 'Thêm xuất món thành công!')
                return redirect('xuat_theo_mon')
            else:
                messages.error(request, 'Lỗi khi thêm xuất món. Vui lòng kiểm tra lại.')
        elif 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            import_type = request.POST.get('import_type')
            try:
                df = pd.read_excel(excel_file)
                if df.empty:
                    messages.error(request, 'File Excel trống. Vui lòng kiểm tra lại.')
                    return redirect('xuat_theo_mon')
                if import_type == 'cong_thuc_mon':
                    required_columns = ['Tên món', 'Nguyên liệu', 'Đơn vị nguyên liệu', 'Định lượng']
                    if not all(col in df.columns for col in required_columns):
                        messages.error(request, f'File thiếu cột bắt buộc. Yêu cầu: {", ".join(required_columns)}.')
                        return redirect('xuat_theo_mon')
                    grouped = df.groupby('Tên món')
                    created_recipes = 0
                    created_details = 0
                    for ten_mon, group in grouped:
                        # Tạo hoặc lấy bản ghi CongThucMon
                        cong_thuc, created = CongThucMon.objects.get_or_create(
                            ten_mon=ten_mon,
                            defaults={}
                        )
                        if created:
                            created_recipes += 1
                        # Lặp qua từng hàng trong nhóm để tạo ChiTietCongThucMon
                        for index, row in group.iterrows():
                            ten_hang_hoa = str(row['Nguyên liệu']).strip() if pd.notnull(row['Nguyên liệu']) else ''
                            don_vi_nguyen_lieu = str(row['Đơn vị nguyên liệu']).strip() if pd.notnull(row['Đơn vị nguyên liệu']) else ''
                            dinh_luong = row['Định lượng']
                            if not all([ten_hang_hoa, don_vi_nguyen_lieu, pd.notnull(dinh_luong)]):
                                messages.error(request, f'Dòng {index + 2}: Thiếu dữ liệu bắt buộc (Nguyên liệu, Đơn vị nguyên liệu, Định lượng).')
                                continue
                            try:
                                # Tạo hoặc lấy HangHoa
                                hang_hoa, created_hang_hoa = HangHoa.objects.get_or_create(
                                    ten_hang_hoa=ten_hang_hoa,
                                    defaults={'don_vi_nguyen_lieu': don_vi_nguyen_lieu}
                                )
                                # Nếu đã tồn tại HangHoa nhưng đơn vị không khớp, cập nhật đơn vị
                                if not created_hang_hoa and hang_hoa.don_vi_nguyen_lieu != don_vi_nguyen_lieu:
                                    hang_hoa.don_vi_nguyen_lieu = don_vi_nguyen_lieu
                                    hang_hoa.save()
                                # Chuyển đổi định lượng sang số
                                dinh_luong = float(dinh_luong) if pd.notna(dinh_luong) else 0
                                if dinh_luong <= 0:
                                    messages.error(request, f'Dòng {index + 2}: Định lượng phải lớn hơn 0.')
                                    continue
                                # Tạo ChiTietCongThucMon
                                ChiTietCongThucMon.objects.update_or_create(
                                    cong_thuc_mon=cong_thuc,
                                    hang_hoa=hang_hoa,
                                    defaults={'dinh_luong': dinh_luong}
                                )
                                created_details += 1
                            except ValueError as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi định dạng số: {str(e)}')
                                continue
                            except Exception as e:
                                messages.error(request, f'Dòng {index + 2}: Lỗi không xác định: {str(e)}')
                                continue
                    messages.success(request, f'Nhập công thức món từ Excel thành công! Đã thêm {created_recipes} món và {created_details} nguyên liệu.')
                    return redirect('xuat_theo_mon')
                elif import_type == 'xuat_mon_fabi':
                    created_entries = 0
                    for index, row in df.iterrows():
                        ngay_xuat = pd.to_datetime(row['ngay_xuat']).date() if pd.notnull(row['ngay_xuat']) else timezone.now().date()
                        ten_mon_instance = CongThucMon.objects.filter(ten_mon=row['ten_mon']).first()
                        if not ten_mon_instance:
                            messages.error(request, f'Món "{row["ten_mon"]}" không tồn tại trong công thức.')
                            continue
                        nhom_mon = row['nhom_mon'] if pd.notnull(row['nhom_mon']) else ''
                        loai_mon = row['loai_mon'] if pd.notnull(row['loai_mon']) else ''
                        don_vi_tinh = row['don_vi_tinh'] if pd.notnull(row['don_vi_tinh']) else ''
                        so_luong = float(row['so_luong']) if pd.notnull(row['so_luong']) else 0
                        xuat_mon = XuatMonTheoFabi.objects.create(
                            ngay_xuat=ngay_xuat,
                            ten_mon=ten_mon_instance,
                            nhom_mon=nhom_mon,
                            loai_mon=loai_mon,
                            don_vi_tinh=don_vi_tinh,
                            so_luong=so_luong
                        )
                        created_entries += 1
                        chi_tiet_list = ChiTietCongThucMon.objects.filter(cong_thuc_mon=ten_mon_instance)
                        for chi_tiet in chi_tiet_list:
                            tong_hop, created = TongHopXuatNguyenLieu.objects.get_or_create(
                                cong_thuc_mon=ten_mon_instance,
                                hang_hoa=chi_tiet.hang_hoa,
                                ngay_xuat=xuat_mon.ngay_xuat,
                                defaults={
                                    'dinh_luong': chi_tiet.dinh_luong,
                                    'so_mon_xuat': so_luong,
                                    'nguyen_lieu_da_xuat': chi_tiet.dinh_luong * so_luong
                                }
                            )
                            if not created:
                                tong_hop.so_mon_xuat += so_luong
                                tong_hop.nguyen_lieu_da_xuat = chi_tiet.dinh_luong * tong_hop.so_mon_xuat
                                tong_hop.save()
                            if abs(tong_hop.nguyen_lieu_da_xuat - (tong_hop.dinh_luong * tong_hop.so_mon_xuat)) > 0.0001:
                                messages.error(request, f'Lỗi toàn vẹn dữ liệu cho {tong_hop.hang_hoa.ten_hang_hoa} ngày {tong_hop.ngay_xuat}')
                    messages.success(request, f'Nhập dữ liệu từ Excel thành công! Đã thêm {created_entries} bản ghi.')
                    return redirect('xuat_theo_mon')
                else:
                    messages.error(request, 'Loại nhập không hợp lệ.')
                    return redirect('xuat_theo_mon')
            except Exception as e:
                messages.error(request, f'Lỗi khi nhập dữ liệu từ Excel: {str(e)}')
    else:
        xuat_mon_list = XuatMonTheoFabi.objects.all()
        tong_hop_list = TongHopXuatNguyenLieu.objects.all()
        cong_thuc_list = CongThucMon.objects.all()

        # Query tổng hợp nguyên liệu
        tong_hop_nguyen_lieu = TongHopXuatNguyenLieu.objects.values(
            'hang_hoa__ten_hang_hoa',
            'hang_hoa__don_vi_nguyen_lieu'
        ).annotate(
            tong_da_xuat=Sum('nguyen_lieu_da_xuat')
        )

        # Áp dụng bộ lọc tìm kiếm nguyên liệu (giữ lại)
        search_nguyen_lieu = request.GET.get('search_nguyen_lieu', '').strip()
        if search_nguyen_lieu:
            tong_hop_nguyen_lieu = tong_hop_nguyen_lieu.filter(hang_hoa__ten_hang_hoa__icontains=search_nguyen_lieu)

        # Chuyển đổi queryset thành list để truyền vào template
        tong_hop_nguyen_lieu_list = [
            {
                'ten_hang_hoa': item['hang_hoa__ten_hang_hoa'],
                'don_vi_nguyen_lieu': item['hang_hoa__don_vi_nguyen_lieu'],
                'tong_da_xuat': item['tong_da_xuat']
            }
            for item in tong_hop_nguyen_lieu
        ]

        hang_hoa_list = HangHoa.objects.all()
        hang_hoa_list_json = json.dumps(
            list(hang_hoa_list.values('pk', 'ten_hang_hoa', 'don_vi_nguyen_lieu')),
            cls=DjangoJSONEncoder
        )
        context = {
            'cong_thuc_form': cong_thuc_form,
            'chi_tiet_form': chi_tiet_form,
            'xuat_mon_form': xuat_mon_form,
            'xuat_mon_list': xuat_mon_list,
            'tong_hop_list': tong_hop_list,
            'hang_hoa_list': hang_hoa_list,
            'hang_hoa_list_json': hang_hoa_list_json,
            'cong_thuc_list': cong_thuc_list,
            'tong_hop_nguyen_lieu_list': tong_hop_nguyen_lieu_list,
            'search_nguyen_lieu': search_nguyen_lieu,
        }
        return render(request, 'home/xuat_theo_mon.html', context)

def edit_ton_kho_le_tan(request, ton_kho_id):
    ton_kho = get_object_or_404(TonKhoLeTan, id=ton_kho_id)
    if request.method == 'POST':
        form = TonKhoLeTanForm(request.POST, instance=ton_kho)
        if form.is_valid():
            ton_kho = form.save(commit=False)
            ton_dau_ngay = tinh_ton_kho_le_tan(ton_kho.hang_hoa.id, ton_kho.ngay_ton)
            ton_kho.ton_dau_ngay = ton_dau_ngay
            ton_kho.ton_cuoi_ngay = form.cleaned_data['ton_cuoi_ngay']
            ton_kho.don_vi_hang_hoa = ton_kho.hang_hoa.don_vi_hang_hoa
            ton_kho.save()
            messages.success(request, 'Sửa tồn kho lễ tân thành công!')
            return redirect(request.POST.get('next', 'quan_ly_hang_hoa'))
        else:
            messages.error(request, 'Lỗi khi sửa tồn kho lễ tân. Vui lòng kiểm tra lại.')
    else:
        form = TonKhoLeTanForm(instance=ton_kho)
    context = {
        'form': form,
        'ton_kho': ton_kho,
        'next': request.GET.get('next', 'quan_ly_hang_hoa'),
    }
    return render(request, 'home/edit_ton_kho_le_tan.html', context)

def delete_ton_kho_le_tan(request, ton_kho_id):
    ton_kho = get_object_or_404(TonKhoLeTan, id=ton_kho_id)
    if request.method == 'POST':
        ton_kho.delete()
        messages.success(request, 'Xóa tồn kho lễ tân thành công!')
        return redirect(request.POST.get('next', 'quan_ly_hang_hoa'))
    return redirect('quan_ly_hang_hoa')

def delete_all_ton_kho_le_tan(request):
    if request.method == 'POST':
        try:
            count = TonKhoLeTan.objects.all().delete()[0]
            messages.success(request, f'Đã xóa {count} bản ghi tồn kho lễ tân thành công!')
        except Exception as e:
            messages.error(request, f'Lỗi khi xóa tất cả bản ghi tồn kho lễ tân: {str(e)}')
        return redirect('quan_ly_hang_hoa')
    return redirect('quan_ly_hang_hoa')

def delete_cong_thuc(request, id):
    if request.method == 'POST':
        cong_thuc = get_object_or_404(CongThucMon, id=id)
        cong_thuc.delete()
        messages.success(request, 'Công thức món đã được xóa thành công!')
        return redirect('xuat_theo_mon')
    return redirect('xuat_theo_mon')

def delete_all_cong_thuc(request):
    if request.method == 'POST':
        CongThucMon.objects.all().delete()
        messages.success(request, 'Tất cả công thức món đã được xóa thành công!')
        return redirect('xuat_theo_mon')
    return redirect('xuat_theo_mon')

def homepage(request):
    today = date.today()
    hang_hoa_form = HangHoaForm()
    nhap_hang_hoa_form = NhapHangHoaForm()
    ton_kho_form = TonKhoHangHoaForm()
    
    # Fetch data
    hang_hoa_list = HangHoa.objects.all()
    ton_kho_hang_hoa_list = TonKhoHangHoa.objects.filter(ngay_ton=today)
    nhap_hang_hoa_list = NhapHangHoa.objects.order_by('-ngay_nhap')[:5]
    ky_ton_kho_list = KyTonKho.objects.all()

    if request.method == 'POST':
        if 'hang_hoa_form' in request.POST:
            hang_hoa_form = HangHoaForm(request.POST)
            if hang_hoa_form.is_valid():
                hang_hoa_form.save()
                messages.success(request, 'Thêm hàng hóa thành công!')
                return redirect('homepage')
            else:
                messages.error(request, 'Có lỗi khi thêm hàng hóa. Vui lòng kiểm tra lại.')
        elif 'nhap_hang_hoa_form' in request.POST:
            nhap_hang_hoa_form = NhapHangHoaForm(request.POST)
            if nhap_hang_hoa_form.is_valid():
                nhap_hang_hoa_form.save()
                messages.success(request, 'Nhập hàng hóa thành công!')
                return redirect('homepage')
            else:
                messages.error(request, 'Có lỗi khi nhập hàng hóa. Vui lòng kiểm tra lại.')
        elif 'ton_kho_hang_hoa_form' in request.POST:
            ton_kho_form = TonKhoHangHoaForm(request.POST)
            if ton_kho_form.is_valid():
                ton_kho_form.save()
                messages.success(request, 'Thêm tồn kho thành công!')
                return redirect('homepage')
            else:
                messages.error(request, 'Có lỗi khi thêm tồn kho. Vui lòng kiểm tra lại.')
        elif 'import_hang_hoa_excel' in request.POST:
            excel_file = request.FILES.get('excel_file')
            if excel_file:
                try:
                    df = pd.read_excel(excel_file)
                    for _, row in df.iterrows():
                        HangHoa.objects.create(
                            ten_hang_hoa=row['ten_hang_hoa'],
                            don_vi_hang_hoa=row['don_vi_hang_hoa'],
                            don_vi_nguyen_lieu=row['don_vi_nguyen_lieu'],
                            dinh_luong=row['dinh_luong']
                        )
                    messages.success(request, 'Nhập hàng hóa từ Excel thành công!')
                except Exception as e:
                    messages.error(request, f'Lỗi khi nhập từ Excel: {str(e)}')
                return redirect('homepage')
            else:
                messages.error(request, 'Vui lòng chọn file Excel.')

    context = {
        'hang_hoa_form': hang_hoa_form,
        'nhap_hang_hoa_form': nhap_hang_hoa_form,
        'ton_kho_form': ton_kho_form,
        'hang_hoa_list': hang_hoa_list,
        'ton_kho_hang_hoa_list': ton_kho_hang_hoa_list,
        'nhap_hang_hoa_list': nhap_hang_hoa_list,
        'ky_ton_kho_list': ky_ton_kho_list,
        'today': today,
    }
    return render(request, 'home/homepage.html', context)

def edit_hang_hoa(request, id):
    hang_hoa = HangHoa.objects.get(id=id)
    if request.method == 'POST':
        form = HangHoaForm(request.POST, instance=hang_hoa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật hàng hóa thành công!')
            next_url = request.GET.get('next', 'homepage')
            return redirect(next_url)
        else:
            messages.error(request, 'Có lỗi khi cập nhật hàng hóa. Vui lòng kiểm tra lại.')
    else:
        form = HangHoaForm(instance=hang_hoa)
    return render(request, 'edit_hang_hoa.html', {'form': form, 'hang_hoa': hang_hoa})

def delete_hang_hoa(request, id):
    if request.method == 'POST':
        try:
            hang_hoa = HangHoa.objects.get(id=id)
            hang_hoa.delete()
            messages.success(request, 'Xóa hàng hóa thành công!')
        except HangHoa.DoesNotExist:
            messages.error(request, 'Hàng hóa không tồn tại.')
        except Exception as e:
            messages.error(request, f'Lỗi khi xóa hàng hóa: {str(e)}')
        next_url = request.POST.get('next', 'homepage')
        return redirect(next_url)
    return redirect('homepage')

def delete_all_hang_hoa(request):
    if request.method == 'POST':
        try:
            count = HangHoa.objects.all().delete()[0]
            messages.success(request, f'Đã xóa {count} hàng hóa thành công!')
        except Exception as e:
            messages.error(request, f'Lỗi khi xóa tất cả hàng hóa: {str(e)}')
        return redirect('homepage')
    return redirect('homepage')

def so_sanh_ton_kho(request, ky_id):
    ky_ton_kho = get_object_or_404(KyTonKho, id=ky_id)
    ton_kho_list = TonKhoHangHoa.objects.filter(ky_ton_kho=ky_ton_kho)
    context = {
        'ky_ton_kho': ky_ton_kho,
        'ton_kho_list': ton_kho_list,
    }
    return render(request, 'home/so_sanh_ton_kho.html', context)

def delete_ton_kho_hang_hoa(request, ton_kho_id):
    ton_kho = get_object_or_404(TonKhoHangHoa, id=ton_kho_id)
    if request.method == 'POST':
        ton_kho.delete()
        messages.success(request, 'Xóa tồn kho thành công!')
        return redirect(request.POST.get('next', 'homepage'))
    return redirect('homepage')

def edit_ton_kho_hang_hoa(request, ton_kho_id):
    ton_kho = get_object_or_404(TonKhoHangHoa, id=ton_kho_id)
    if request.method == 'POST':
        form = TonKhoHangHoaForm(request.POST, instance=ton_kho)
        if form.is_valid():
            ton_kho = form.save(commit=False)
            ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(ton_kho.hang_hoa.id, ton_kho.ngay_ton)
            ton_kho.ton_dau_ngay = ton_dau_ngay
            ton_kho.ton_cuoi_ngay = form.cleaned_data['ton_cuoi_ngay']
            ton_kho.don_vi_hang_hoa = ton_kho.hang_hoa.don_vi_hang_hoa
            ton_kho.save()
            messages.success(request, 'Sửa tồn kho thành công!')
            return redirect(request.POST.get('next', 'homepage'))
        else:
            messages.error(request, 'Lỗi khi sửa tồn kho. Vui lòng kiểm tra lại.')
    else:
        form = TonKhoHangHoaForm(instance=ton_kho)
    context = {
        'form': form,
        'ton_kho': ton_kho,
        'next': request.GET.get('next', 'homepage'),
    }
    return render(request, 'home/edit_ton_kho_hang_hoa.html', context)

def quan_ly_ton_kho(request):
    form = TonKhoHangHoaForm()
    import_form = TonKhoHangHoaImportForm()
    filter_form = TonKhoHangHoaFilterForm(request.GET or None)

    ton_kho_list = TonKhoHangHoa.objects.all()

    search_query = request.GET.get('search', '')
    if search_query:
        ton_kho_list = ton_kho_list.filter(hang_hoa__ten_hang_hoa__icontains=search_query)

    if filter_form.is_valid():
        ngay_bat_dau = filter_form.cleaned_data.get('ngay_bat_dau')
        ngay_ket_thuc = filter_form.cleaned_data.get('ngay_ket_thuc')
        if ngay_bat_dau:
            ton_kho_list = ton_kho_list.filter(ngay_ton__gte=ngay_bat_dau)
        if ngay_ket_thuc:
            ton_kho_list = ton_kho_list.filter(ngay_ton__lte=ngay_ket_thuc)

    paginator = Paginator(ton_kho_list.order_by('ngay_ton'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        if 'excel_import' in request.POST:
            import_form = TonKhoHangHoaImportForm(request.POST, request.FILES)
            if import_form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    df = pd.read_excel(excel_file)
                    for index, row in df.iterrows():
                        ten_hang_hoa = row.get('Tên Hàng Hóa')
                        ngay_ton = row.get('Ngày Tồn')
                        ton_cuoi_ngay = row.get('Tồn Cuối Ngày')

                        if ten_hang_hoa and ngay_ton and ton_cuoi_ngay is not None:
                            try:
                                ngay_ton = pd.to_datetime(ngay_ton).date()
                                hang_hoa = HangHoa.objects.get(ten_hang_hoa=ten_hang_hoa)
                                ton_dau_ngay, _ = tinh_ton_kho(hang_hoa.id, ngay_ton)
                                TonKhoHangHoa.objects.update_or_create(
                                    hang_hoa=hang_hoa,
                                    ngay_ton=ngay_ton,
                                    defaults={
                                        'ton_dau_ngay': ton_dau_ngay,
                                        'ton_cuoi_ngay': float(ton_cuoi_ngay),
                                        'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa
                                    }
                                )
                            except HangHoa.DoesNotExist:
                                messages.error(request, f'Hàng hóa "{ten_hang_hoa}" không tồn tại tại dòng {index + 2}')
                                continue
                            except ValueError:
                                messages.error(request, f'Lỗi định dạng ngày tại dòng {index + 2}')
                                continue
                    messages.success(request, 'Nhập dữ liệu từ Excel thành công!')
                    return redirect('quan_ly_ton_kho')
                except Exception as e:
                    messages.error(request, f'Lỗi khi nhập file Excel: {str(e)}')
            else:
                messages.error(request, 'Lỗi khi nhập file Excel. Vui lòng kiểm tra định dạng.')
        else:
            form = TonKhoHangHoaForm(request.POST)
            if form.is_valid():
                ton_kho = form.save(commit=False)
                ton_dau_ngay, ton_cuoi_ngay = tinh_ton_kho(ton_kho.hang_hoa.id, ton_kho.ngay_ton)
                ton_kho.ton_dau_ngay = ton_dau_ngay
                ton_kho.ton_cuoi_ngay = form.cleaned_data['ton_cuoi_ngay']
                ton_kho.don_vi_hang_hoa = ton_kho.hang_hoa.don_vi_hang_hoa
                ton_kho.save()
                messages.success(request, 'Thêm tồn kho thành công!')
                return redirect('quan_ly_ton_kho')
            else:
                messages.error(request, 'Lỗi khi thêm tồn kho. Vui lòng kiểm tra lại.')

    context = {
        'form': form,
        'import_form': import_form,
        'filter_form': filter_form,
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'home/quan_ly_ton_kho.html', context)


def so_sanh_fabi(request):
    # Lấy tham số tìm kiếm và lọc
    ten_hang_hoa = request.GET.get('ten_hang_hoa', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Chuyển đổi ngày nếu có
    try:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
    except ValueError:
        date_from = None
        date_to = None

    # Lấy dữ liệu từ TongHopXuatNguyenLieu (Nguyên Liệu Xuất - Fabi)
    tong_hop_query = TongHopXuatNguyenLieu.objects.all()
    if ten_hang_hoa:
        tong_hop_query = tong_hop_query.filter(hang_hoa__ten_hang_hoa__icontains=ten_hang_hoa)
    if date_from:
        tong_hop_query = tong_hop_query.filter(ngay_xuat__gte=date_from)
    if date_to:
        tong_hop_query = tong_hop_query.filter(ngay_xuat__lte=date_to)

    tong_hop_data = tong_hop_query.values('hang_hoa__ten_hang_hoa').annotate(
        tong_da_xuat=Sum('nguyen_lieu_da_xuat')
    )

    # Lấy dữ liệu lượng dùng từ TonKhoHangHoa
    luong_dung_query = TonKhoHangHoa.objects.all()
    if ten_hang_hoa:
        luong_dung_query = luong_dung_query.filter(hang_hoa__ten_hang_hoa__icontains=ten_hang_hoa)
    if date_from:
        luong_dung_query = luong_dung_query.filter(ngay_ton__gte=date_from)
    if date_to:
        luong_dung_query = luong_dung_query.filter(ngay_ton__lte=date_to)

    luong_dung_data = luong_dung_query.values('hang_hoa__ten_hang_hoa').annotate(
        luong_dung_tong=Sum(F('ton_dau_ngay') - F('ton_cuoi_ngay'))
    )

    # Kết hợp dữ liệu
    so_sanh_list = []
    hang_hoa_list = HangHoa.objects.all()
    if ten_hang_hoa:
        hang_hoa_list = hang_hoa_list.filter(ten_hang_hoa__icontains=ten_hang_hoa)

    for hang_hoa in hang_hoa_list:
        fabi_data = next((item for item in tong_hop_data if item['hang_hoa__ten_hang_hoa'] == hang_hoa.ten_hang_hoa), None)
        nguyen_lieu_da_xuat = fabi_data['tong_da_xuat'] if fabi_data else 0

        luong_dung = next((item for item in luong_dung_data if item['hang_hoa__ten_hang_hoa'] == hang_hoa.ten_hang_hoa), None)
        so_luong_ton = abs(luong_dung['luong_dung_tong']) if luong_dung and luong_dung['luong_dung_tong'] is not None else 0

        chenh_lech = so_luong_ton - nguyen_lieu_da_xuat

        if nguyen_lieu_da_xuat or so_luong_ton:
            so_sanh_list.append({
                'hang_hoa': hang_hoa,
                'nguyen_lieu_da_xuat': nguyen_lieu_da_xuat,
                'so_luong_ton': so_luong_ton,
                'chenh_lech': chenh_lech,
                'date_from': date_from or timezone.now().date(),
                'date_to': date_to or timezone.now().date(),
            })

    context = {
        'so_sanh_list': so_sanh_list,
        'ten_hang_hoa': ten_hang_hoa,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'home/so_sanh_fabi.html', context)



def delete_nhap_hang_hoa(request, id):
    if request.method == 'POST':
        nhap_hang_hoa = get_object_or_404(NhapHangHoa, id=id)
        nhap_hang_hoa.delete()
        messages.success(request, 'Bản ghi nhập hàng đã được xóa thành công!')
        return redirect('quan_ly_hang_hoa')  # Chuyển hướng về trang quản lý hàng hóa
    return redirect('home/quan_ly_hang_hoa.html')

def delete_xuat_mon(request, id):
    if request.method == 'POST':
        xuat_mon = get_object_or_404(XuatMonTheoFabi, id=id)
        xuat_mon.delete()
        messages.success(request, 'Bản ghi xuất món đã được xóa thành công!')
        return redirect('xuat_theo_mon')
    return redirect('xuat_theo_mon')

def delete_all_xuat_mon(request):
    if request.method == 'POST':
        XuatMonTheoFabi.objects.all().delete()
        messages.success(request, 'Tất cả bản ghi xuất món đã được xóa thành công!')
        return redirect('xuat_theo_mon')
    return redirect('xuat_theo_mon')

def delete_tong_hop(request, id):
    if request.method == 'POST':
        tong_hop = get_object_or_404(TongHopXuatNguyenLieu, id=id)
        tong_hop.delete()
        messages.success(request, 'Bản ghi tổng hợp xuất nguyên liệu đã được xóa thành công!')
        return redirect('xuat_theo_mon')
    return redirect('xuat_theo_mon')

def delete_all_tong_hop(request):
    if request.method == 'POST':
        TongHopXuatNguyenLieu.objects.all().delete()
        messages.success(request, 'Tất cả bản ghi tổng hợp xuất nguyên liệu đã được xóa thành công!')
        return redirect('xuat_theo_mon')
    return redirect('xuat_theo_mon')


def delete_all_nhap_hang(request):
    if request.method == 'POST':
        try:
            count = NhapHangHoa.objects.all().delete()[0]
            messages.success(request, f'Đã xóa {count} bản ghi nhập hàng thành công!')
        except Exception as e:
            messages.error(request, f'Lỗi khi xóa tất cả bản ghi nhập hàng: {str(e)}')
        return redirect('quan_ly_hang_hoa')
    return redirect('quan_ly_hang_hoa')


def delete_ton_kho_by_date_range(request):
    if request.method == 'POST':
        ngay_bat_dau = request.POST.get('ngay_bat_dau')
        ngay_ket_thuc = request.POST.get('ngay_ket_thuc')
        try:
            ngay_bat_dau = datetime.strptime(ngay_bat_dau, '%Y-%m-%d').date()
            ngay_ket_thuc = datetime.strptime(ngay_ket_thuc, '%Y-%m-%d').date()
            if ngay_bat_dau > ngay_ket_thuc:
                messages.error(request, 'Ngày bắt đầu phải nhỏ hơn hoặc bằng ngày kết thúc.')
                return redirect('quan_ly_hang_hoa')
            
            count = TonKhoHangHoa.objects.filter(
                ngay_ton__range=[ngay_bat_dau, ngay_ket_thuc]
            ).delete()[0]
            messages.success(request, f'Đã xóa {count} bản ghi tồn kho trong khoảng từ {ngay_bat_dau} đến {ngay_ket_thuc} thành công!')
        except ValueError:
            messages.error(request, 'Định dạng ngày không hợp lệ.')
        except Exception as e:
            messages.error(request, f'Lỗi khi xóa bản ghi tồn kho: {str(e)}')
        return redirect('quan_ly_hang_hoa')
    return redirect('quan_ly_hang_hoa')

@require_GET
def get_ton_kho_le_tan_data(request):
    """API lấy dữ liệu tồn kho lễ tân theo ngày"""
    ngay_ton_str = request.GET.get('ngay_ton')
    try:
        ngay_ton = datetime.strptime(ngay_ton_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Ngày không hợp lệ'}, status=400)
    
    hang_hoa_list = HangHoa.objects.all()
    data = []
    
    for hang_hoa in hang_hoa_list:
        ton_dau_ngay = tinh_ton_kho_le_tan(hang_hoa.id, ngay_ton)
        
        existing_ton_kho = TonKhoLeTan.objects.filter(
            hang_hoa=hang_hoa,
            ngay_ton=ngay_ton
        ).first()
        
        data.append({
            'hang_hoa_id': hang_hoa.id,
            'ten_hang_hoa': hang_hoa.ten_hang_hoa,
            'ton_dau_ngay': float(ton_dau_ngay),
            'ton_cuoi_ngay': float(existing_ton_kho.ton_cuoi_ngay) if existing_ton_kho else None,
            'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa or 'Không xác định'
        })
    
    return JsonResponse({'success': True, 'data': data})


@require_GET
def get_hang_hoa_list(request):
    """API lấy danh sách hàng hóa"""
    hang_hoa_list = HangHoa.objects.all().values(
        'id', 'ten_hang_hoa', 'don_vi_nguyen_lieu'
    )
    
    ngay_ton_str = request.GET.get('ngay_ton')
    try:
        ngay_ton = datetime.strptime(ngay_ton_str, '%Y-%m-%d').date()
    except ValueError:
        from datetime import date
        ngay_ton = date.today()
    
    data = []
    for hh in hang_hoa_list:
        ton_dau_ngay, _ = tinh_ton_kho(hh['id'], ngay_ton)
        data.append({
            'id': hh['id'],
            'ten_hang_hoa': hh['ten_hang_hoa'],
            'don_vi_nguyen_lieu': hh['don_vi_nguyen_lieu'] or 'Không xác định',
            'ton_dau_ngay': float(ton_dau_ngay)
        })
    
    return JsonResponse(data, safe=False)

# THÊM VÀO CUỐI views.py
@require_GET
def get_hang_hoa_detail(request, id):
    """API lấy chi tiết một hàng hóa"""
    try:
        hang_hoa = HangHoa.objects.get(id=id)
        
        # Tính tồn đầu ngày cho ngày được chọn
        ngay_ton_str = request.GET.get('ngay_ton')
        try:
            ngay_ton = datetime.strptime(ngay_ton_str, '%Y-%m-%d').date()
        except ValueError:
            from datetime import date
            ngay_ton = date.today()
        
        ton_dau_ngay, _ = tinh_ton_kho(hang_hoa.id, ngay_ton)
        
        data = {
            'success': True,
            'data': {
                'id': hang_hoa.id,
                'ten_hang_hoa': hang_hoa.ten_hang_hoa,
                'don_vi_nguyen_lieu': hang_hoa.don_vi_nguyen_lieu or 'Không xác định',
                'don_vi_hang_hoa': hang_hoa.don_vi_hang_hoa or 'Không xác định',
                'dinh_luong': float(hang_hoa.dinh_luong) if hang_hoa.dinh_luong else 1,
                'ton_dau_ngay': float(ton_dau_ngay)
            }
        }
        return JsonResponse(data)
        
    except HangHoa.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hàng hóa không tồn tại'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)